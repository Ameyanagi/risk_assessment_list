from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from risk_assessment_list.ghs import (
    GHS_CLASS_SPECS,
    classification_state,
    derive_pictograms,
    extract_category_codes,
)
from risk_assessment_list.normalize import (
    extract_cas_rns,
    normalize_cas,
    normalize_text,
    parse_float,
)
from risk_assessment_list.synonym_database import build_synonym_database
from risk_assessment_list.synonyms import generate_synonyms, normalize_synonym_text

REFERENCE_DIR = ROOT / "reference"
MANIFEST_PATH = REFERENCE_DIR / "manifest.json"
BUILD_DIR = ROOT / "build"
OUTPUT_DB = ROOT / "src/risk_assessment_list/data/ra.sqlite3"
SYNONYM_DB_OUTPUT = REFERENCE_DIR / "generated_synonyms.sqlite3"
REVIEWED_SYNONYM_CSV = REFERENCE_DIR / "reviewed_synonyms.csv"
TEMP_DB = BUILD_DIR / "ra.sqlite3.tmp"
SCHEMA_VERSION = 2

LEGAL_SOURCE_FILES = ["001168179.xlsx", "001474394.xlsx"]
GHS_SOURCE_FILE = "list_nite_all.xlsx"
SNAPSHOT_BASELINES = {
    "8ddba2f69b917a4c5386d37cc062bec2a6792e9f4f8faef3c3a08dd749b9d3ca": {
        "raw_legal_rows_total": 4953,
        "raw_legal_rows_data": 4889,
        "raw_ghs_rows_total": 3418,
        "raw_ghs_rows_data": 3417,
        "legal_obligations": 4889,
        "legal_obligation_cas": 4951,
        "ghs_entries": 3417,
        "ghs_entry_cas": 3453,
        "ghs_hazard_classes": 35,
    },
}


def require_lastrowid(cursor: sqlite3.Cursor) -> int:
    lastrowid = cursor.lastrowid
    if lastrowid is None:
        raise RuntimeError("sqlite cursor did not report lastrowid")
    return int(lastrowid)


@dataclass(frozen=True)
class LegalRecord:
    source_file_id: int
    source_sheet: str
    source_row: int
    substance_name: str
    english_name: str
    cas_text: str
    cas_rns: tuple[str, ...]
    section_title: str
    section_number: str
    list_index: str
    label_threshold: float | None
    sds_threshold: float | None
    remarks: str
    source_list_effective_date: str | None
    raw_effective_date: str


@dataclass(frozen=True)
class GHSRecord:
    source_file_id: int
    source_sheet: str
    source_row: int
    substance_name: str
    english_name: str
    cas_text: str
    cas_rns: tuple[str, ...]
    ghs_result_id: str
    hazard_classes: dict[str, str]
    model_label_url: str | None
    model_sds_url: str | None


class Canonicalizer:
    def __init__(self, connection: sqlite3.Connection, fts_enabled: bool) -> None:
        self.connection = connection
        self.fts_enabled = fts_enabled
        self.cas_to_substance_id: dict[str, int] = {}
        self.name_to_substance_id: dict[str, int] = {}
        self.identifier_keys: set[tuple[int, str, str]] = set()
        self.alias_keys: set[tuple[int, str]] = set()

    def get_or_create_substance(
        self,
        *,
        substance_name: str,
        english_name: str,
        cas_rns: tuple[str, ...],
    ) -> int:
        substance_id = self._lookup_substance_id(
            cas_rns=cas_rns,
            name_keys=self._name_keys(substance_name, english_name),
        )
        if substance_id is None:
            cursor = self.connection.execute(
                """
                insert into substances (
                    canonical_name,
                    canonical_name_normalized,
                    canonical_english_name,
                    canonical_english_name_normalized,
                    canonical_cas,
                    substance_kind
                ) values (?, ?, ?, ?, ?, ?)
                """,
                (
                    substance_name,
                    normalize_text(substance_name),
                    english_name or None,
                    normalize_text(english_name) if english_name else None,
                    cas_rns[0] if cas_rns else None,
                    "cas_linked" if cas_rns else "name_only",
                ),
            )
            substance_id = require_lastrowid(cursor)
        else:
            self._refresh_substance(
                substance_id=substance_id,
                substance_name=substance_name,
                english_name=english_name,
                cas_rns=cas_rns,
            )

        for cas_rn in cas_rns:
            self.cas_to_substance_id[normalize_cas(cas_rn)] = substance_id
        for name_key in self._name_keys(substance_name, english_name):
            self.name_to_substance_id[name_key] = substance_id

        self._add_identifiers(substance_id, substance_name, english_name, cas_rns)
        self._add_aliases(substance_id, substance_name, english_name, cas_rns)
        return substance_id

    def _lookup_substance_id(
        self,
        *,
        cas_rns: tuple[str, ...],
        name_keys: tuple[str, ...],
    ) -> int | None:
        for cas_rn in cas_rns:
            substance_id = self.cas_to_substance_id.get(normalize_cas(cas_rn))
            if substance_id is not None:
                return substance_id
        for name_key in name_keys:
            substance_id = self.name_to_substance_id.get(name_key)
            if substance_id is not None:
                return substance_id
        return None

    def _refresh_substance(
        self,
        *,
        substance_id: int,
        substance_name: str,
        english_name: str,
        cas_rns: tuple[str, ...],
    ) -> None:
        self.connection.execute(
            """
            update substances
            set canonical_name = coalesce(canonical_name, ?),
                canonical_name_normalized = coalesce(canonical_name_normalized, ?),
                canonical_english_name = coalesce(canonical_english_name, ?),
                canonical_english_name_normalized = coalesce(canonical_english_name_normalized, ?),
                canonical_cas = coalesce(canonical_cas, ?)
            where id = ?
            """,
            (
                substance_name,
                normalize_text(substance_name),
                english_name or None,
                normalize_text(english_name) if english_name else None,
                cas_rns[0] if cas_rns else None,
                substance_id,
            ),
        )

    def _add_identifiers(
        self,
        substance_id: int,
        substance_name: str,
        english_name: str,
        cas_rns: tuple[str, ...],
    ) -> None:
        for cas_rn in cas_rns:
            self._insert_identifier(
                substance_id=substance_id,
                identifier_type="cas",
                value_raw=cas_rn,
                value_normalized=normalize_cas(cas_rn),
                exact_match_allowed=True,
                confidence="high",
                source="downloaded_source",
                is_primary=int(cas_rn == cas_rns[0]),
            )
        self._insert_identifier(
            substance_id=substance_id,
            identifier_type="name_ja",
            value_raw=substance_name,
            value_normalized=normalize_text(substance_name),
            exact_match_allowed=True,
            confidence="high",
            source="downloaded_source",
            is_primary=1,
        )
        if english_name:
            self._insert_identifier(
                substance_id=substance_id,
                identifier_type="name_en",
                value_raw=english_name,
                value_normalized=normalize_text(english_name),
                exact_match_allowed=True,
                confidence="high",
                source="downloaded_source",
                is_primary=1,
            )

    def _insert_identifier(
        self,
        *,
        substance_id: int,
        identifier_type: str,
        value_raw: str,
        value_normalized: str,
        exact_match_allowed: bool,
        confidence: str,
        source: str,
        is_primary: int,
    ) -> None:
        key = (substance_id, identifier_type, value_normalized)
        if not value_normalized or key in self.identifier_keys:
            return
        self.identifier_keys.add(key)
        self.connection.execute(
            """
            insert into substance_identifiers (
                substance_id,
                identifier_type,
                value_raw,
                value_normalized,
                exact_match_allowed,
                confidence,
                source,
                is_primary
            ) values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                substance_id,
                identifier_type,
                value_raw,
                value_normalized,
                int(exact_match_allowed),
                confidence,
                source,
                is_primary,
            ),
        )

    def _add_aliases(
        self,
        substance_id: int,
        substance_name: str,
        english_name: str,
        cas_rns: tuple[str, ...],
    ) -> None:
        for cas_rn in cas_rns:
            self.add_alias(
                substance_id=substance_id,
                alias_raw=cas_rn,
                alias_type="cas",
                confidence="high",
                exact_match_allowed=True,
            )
        for raw_value, alias_type in (
            (substance_name, "canonical_name"),
            (english_name, "canonical_english_name"),
        ):
            if not raw_value:
                continue
            self.add_alias(
                substance_id=substance_id,
                alias_raw=raw_value,
                alias_type=alias_type,
                confidence="high",
                exact_match_allowed=True,
            )
            for synonym in generate_synonyms(raw_value):
                if normalize_synonym_text(synonym) == normalize_synonym_text(raw_value):
                    continue
                self.add_alias(
                    substance_id=substance_id,
                    alias_raw=synonym,
                    alias_type="generated_synonym",
                    confidence=_alias_confidence(synonym),
                    exact_match_allowed=False,
                )

    def add_alias(
        self,
        *,
        substance_id: int,
        alias_raw: str,
        alias_type: str,
        confidence: str,
        exact_match_allowed: bool,
    ) -> None:
        alias_normalized = normalize_synonym_text(alias_raw)
        if not _should_keep_alias(alias_raw, alias_normalized):
            return
        key = (substance_id, alias_normalized)
        if key in self.alias_keys:
            return
        cursor = self.connection.execute(
            """
            insert into substance_aliases (
                substance_id,
                alias_raw,
                alias_normalized,
                alias_type,
                confidence,
                exact_match_allowed
            ) values (?, ?, ?, ?, ?, ?)
            """,
            (
                substance_id,
                alias_raw,
                alias_normalized,
                alias_type,
                confidence,
                int(exact_match_allowed),
            ),
        )
        self.alias_keys.add(key)
        if self.fts_enabled:
            self.connection.execute(
                """
                insert into substance_alias_fts (alias_normalized, substance_id, alias_id)
                values (?, ?, ?)
                """,
                (alias_normalized, substance_id, require_lastrowid(cursor)),
            )

    def resolve_substance_id(
        self,
        *,
        canonical_name: str,
        canonical_english_name: str,
        canonical_cas: str,
    ) -> int | None:
        if canonical_cas:
            substance_id = self.cas_to_substance_id.get(normalize_cas(canonical_cas))
            if substance_id is not None:
                return substance_id
        if canonical_name:
            substance_id = self.name_to_substance_id.get(normalize_text(canonical_name))
            if substance_id is not None:
                return substance_id
        if canonical_english_name:
            substance_id = self.name_to_substance_id.get(
                normalize_text(canonical_english_name)
            )
            if substance_id is not None:
                return substance_id
        return None

    @staticmethod
    def _name_keys(substance_name: str, english_name: str) -> tuple[str, ...]:
        keys = []
        if substance_name:
            keys.append(normalize_text(substance_name))
        if english_name:
            keys.append(normalize_text(english_name))
        return tuple(key for key in keys if key)


def main() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    manifest_sha = hashlib.sha256(MANIFEST_PATH.read_bytes()).hexdigest()
    source_snapshot_sha = compute_source_snapshot_sha(manifest["sources"])
    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DB.parent.mkdir(parents=True, exist_ok=True)
    if TEMP_DB.exists():
        TEMP_DB.unlink()

    connection = sqlite3.connect(TEMP_DB)
    connection.execute("pragma foreign_keys = on")
    connection.execute("pragma temp_store = memory")
    connection.execute(f"pragma user_version = {SCHEMA_VERSION}")

    try:
        fts_enabled = create_schema(connection)
        source_file_ids = load_sources(connection, manifest["sources"])
        raw_legal_row_count = 0
        raw_ghs_row_count = 0
        legal_records: list[LegalRecord] = []

        for filename in LEGAL_SOURCE_FILES:
            rows, records = stage_legal_workbook(
                connection,
                source_file_id=source_file_ids[filename],
                path=REFERENCE_DIR / filename,
            )
            raw_legal_row_count += rows
            legal_records.extend(records)

        raw_rows, ghs_records = stage_ghs_workbook(
            connection,
            source_file_id=source_file_ids[GHS_SOURCE_FILE],
            path=REFERENCE_DIR / GHS_SOURCE_FILE,
        )
        raw_ghs_row_count += raw_rows

        hazard_class_ids = seed_ghs_hazard_classes(connection)
        canonicalizer = Canonicalizer(connection, fts_enabled=fts_enabled)

        normalize_legal_records(connection, canonicalizer, legal_records)
        normalize_ghs_records(connection, canonicalizer, ghs_records, hazard_class_ids)
        load_reviewed_synonyms(canonicalizer, REVIEWED_SYNONYM_CSV)

        validate_snapshot(
            connection,
            source_snapshot_sha=source_snapshot_sha,
            expected_source_count=len(manifest["sources"]),
            expected_ghs_classes_per_entry=len(GHS_CLASS_SPECS),
        )

        write_build_meta(
            connection,
            manifest_sha=manifest_sha,
            source_file_count=len(manifest["sources"]),
            raw_legal_row_count=raw_legal_row_count,
            raw_ghs_row_count=raw_ghs_row_count,
            fts_enabled=fts_enabled,
        )

        create_indexes(connection)
        connection.commit()
        connection.execute("pragma optimize")
        connection.execute("vacuum")
    finally:
        connection.close()

    os.replace(TEMP_DB, OUTPUT_DB)
    build_synonym_database(
        OUTPUT_DB,
        SYNONYM_DB_OUTPUT,
        source_manifest_sha=manifest_sha,
    )


def create_schema(connection: sqlite3.Connection) -> bool:
    connection.executescript(
        """
        create table source_files (
            id integer primary key,
            source_key text not null unique,
            url text not null,
            filename text not null unique,
            fetched_at text not null,
            last_modified text,
            etag text,
            content_type text,
            sha256 text not null,
            not_modified integer not null default 0
        );

        create table build_meta (
            id integer primary key check (id = 1),
            schema_version integer not null,
            built_at text not null,
            source_manifest_sha256 text not null,
            source_file_count integer not null,
            raw_legal_row_count integer not null,
            raw_ghs_row_count integer not null,
            legal_obligation_count integer not null,
            ghs_entry_count integer not null,
            substance_count integer not null,
            fts_enabled integer not null
        );

        create table raw_legal_rows (
            id integer primary key,
            source_file_id integer not null references source_files(id),
            source_sheet text not null,
            row_number integer not null,
            row_kind text not null,
            cells_json text not null,
            parsed_ok integer not null
        );

        create table raw_ghs_rows (
            id integer primary key,
            source_file_id integer not null references source_files(id),
            source_sheet text not null,
            row_number integer not null,
            row_kind text not null,
            cells_json text not null,
            parsed_ok integer not null
        );

        create table substances (
            id integer primary key,
            canonical_name text not null,
            canonical_name_normalized text not null,
            canonical_english_name text,
            canonical_english_name_normalized text,
            canonical_cas text,
            substance_kind text not null
        );

        create table substance_identifiers (
            id integer primary key,
            substance_id integer not null references substances(id) on delete cascade,
            identifier_type text not null,
            value_raw text not null,
            value_normalized text not null,
            exact_match_allowed integer not null,
            confidence text not null,
            source text not null,
            is_primary integer not null default 0,
            unique (substance_id, identifier_type, value_normalized)
        );

        create table substance_aliases (
            id integer primary key,
            substance_id integer not null references substances(id) on delete cascade,
            alias_raw text not null,
            alias_normalized text not null,
            alias_type text not null,
            confidence text not null,
            exact_match_allowed integer not null,
            unique (substance_id, alias_normalized)
        );

        create table legal_obligations (
            id integer primary key,
            substance_id integer not null references substances(id) on delete cascade,
            source_file_id integer not null references source_files(id),
            source_sheet text not null,
            source_row integer not null,
            substance_name text not null,
            english_name text,
            cas_text text not null,
            section_title text not null,
            section_number text,
            list_index text,
            label_threshold real,
            sds_threshold real,
            remarks text,
            source_list_effective_date text,
            raw_effective_date text
        );

        create table legal_obligation_cas (
            legal_obligation_id integer not null references legal_obligations(id) on delete cascade,
            cas_rn text not null,
            cas_normalized text not null,
            primary key (legal_obligation_id, cas_rn)
        );

        create table ghs_entries (
            id integer primary key,
            substance_id integer not null references substances(id) on delete cascade,
            source_file_id integer not null references source_files(id),
            source_sheet text not null,
            source_row integer not null,
            substance_name text not null,
            english_name text,
            cas_text text not null,
            ghs_result_id text not null,
            model_label_url text,
            model_sds_url text
        );

        create table ghs_entry_cas (
            ghs_entry_id integer not null references ghs_entries(id) on delete cascade,
            cas_rn text not null,
            cas_normalized text not null,
            primary key (ghs_entry_id, cas_rn)
        );

        create table ghs_hazard_classes (
            id integer primary key,
            code text not null unique,
            label_ja text not null unique,
            sort_order integer not null unique
        );

        create table ghs_classifications (
            ghs_entry_id integer not null references ghs_entries(id) on delete cascade,
            hazard_class_id integer not null references ghs_hazard_classes(id),
            raw_result text not null,
            classification_state text not null,
            category_code text,
            is_assigned integer not null,
            primary key (ghs_entry_id, hazard_class_id)
        );

        create table ghs_pictograms (
            ghs_entry_id integer not null references ghs_entries(id) on delete cascade,
            pictogram_code text not null,
            primary key (ghs_entry_id, pictogram_code)
        );
        """
    )

    try:
        connection.execute(
            """
            create virtual table substance_alias_fts using fts5(
                alias_normalized,
                substance_id unindexed,
                alias_id unindexed,
                tokenize = 'unicode61',
                prefix = '2 3 4 5 6 7 8 9 10 11 12'
            )
            """
        )
        return True
    except sqlite3.OperationalError:
        return False


def create_indexes(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        create index idx_substances_canonical_name on substances(canonical_name_normalized);
        create index idx_substances_canonical_cas on substances(canonical_cas);
        create index idx_substance_identifiers_lookup on substance_identifiers(value_normalized, identifier_type);
        create index idx_substance_aliases_lookup on substance_aliases(alias_normalized);
        create index idx_legal_obligations_substance on legal_obligations(substance_id);
        create index idx_legal_obligation_cas_lookup on legal_obligation_cas(cas_normalized);
        create index idx_ghs_entries_substance on ghs_entries(substance_id);
        create index idx_ghs_entry_cas_lookup on ghs_entry_cas(cas_normalized);
        create index idx_ghs_classifications_assigned on ghs_classifications(ghs_entry_id, is_assigned);
        """
    )


def load_sources(
    connection: sqlite3.Connection, sources: list[dict[str, str]]
) -> dict[str, int]:
    source_ids: dict[str, int] = {}
    for source in sources:
        cursor = connection.execute(
            """
            insert into source_files (
                source_key,
                url,
                filename,
                fetched_at,
                last_modified,
                etag,
                content_type,
                sha256,
                not_modified
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source["key"],
                source["url"],
                source["filename"],
                source["fetched_at"],
                source.get("last_modified", ""),
                source.get("etag", ""),
                source.get("content_type", ""),
                source["sha256"],
                int(bool(source.get("not_modified", False))),
            ),
        )
        source_ids[source["filename"]] = require_lastrowid(cursor)
    return source_ids


def stage_legal_workbook(
    connection: sqlite3.Connection,
    *,
    source_file_id: int,
    path: Path,
) -> tuple[int, list[LegalRecord]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    raw_row_count = 0
    records: list[LegalRecord] = []

    for sheet in workbook.worksheets:
        sheet_effective_date = infer_sheet_effective_date(sheet.title)
        current_section = ""
        current_headers: dict[str, int] | None = None

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [normalize_cell(value) for value in row]
            if not any(values):
                continue

            row_kind = "other"
            parsed_ok = 0
            header_map = detect_legal_header(values)

            if is_section_row(values):
                row_kind = "section"
                current_section = first_non_empty(values)
                current_headers = None
            elif header_map:
                row_kind = "header"
                current_headers = header_map
            elif current_headers and is_legal_data_row(values, current_headers):
                row_kind = "data"
                parsed_ok = 1
                records.append(
                    LegalRecord(
                        source_file_id=source_file_id,
                        source_sheet=sheet.title,
                        source_row=row_index,
                        substance_name=values[current_headers["name"]],
                        english_name=values[current_headers.get("english_name", -1)]
                        if "english_name" in current_headers
                        else "",
                        cas_text=values[current_headers["cas"]],
                        cas_rns=tuple(extract_cas_rns(values[current_headers["cas"]])),
                        section_title=current_section,
                        section_number=values[current_headers.get("section_number", -1)]
                        if "section_number" in current_headers
                        else "",
                        list_index=values[current_headers.get("list_index", -1)]
                        if "list_index" in current_headers
                        else "",
                        label_threshold=parse_float(
                            values[current_headers["label_threshold"]]
                        ),
                        sds_threshold=parse_float(
                            values[current_headers["sds_threshold"]]
                        ),
                        remarks=values[current_headers.get("remarks", -1)]
                        if "remarks" in current_headers
                        else "",
                        source_list_effective_date=sheet_effective_date,
                        raw_effective_date=values[
                            current_headers.get("effective_date", -1)
                        ]
                        if "effective_date" in current_headers
                        else "",
                    )
                )

            connection.execute(
                """
                insert into raw_legal_rows (
                    source_file_id,
                    source_sheet,
                    row_number,
                    row_kind,
                    cells_json,
                    parsed_ok
                ) values (?, ?, ?, ?, ?, ?)
                """,
                (
                    source_file_id,
                    sheet.title,
                    row_index,
                    row_kind,
                    json.dumps(values, ensure_ascii=False),
                    parsed_ok,
                ),
            )
            raw_row_count += 1

    return raw_row_count, records


def stage_ghs_workbook(
    connection: sqlite3.Connection,
    *,
    source_file_id: int,
    path: Path,
) -> tuple[int, list[GHSRecord]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    raw_row_count = 0
    records: list[GHSRecord] = []
    headers: list[str] = []
    positions: dict[str, int] = {}

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [normalize_cell(value) for value in row]
        if not any(values):
            continue

        row_kind = "data"
        parsed_ok = 1
        if row_index == 1:
            headers = values
            positions = {header: index for index, header in enumerate(headers)}
            row_kind = "header"
            parsed_ok = 0
        else:
            hazard_classes = {
                label: values[positions[label]] for _, label in GHS_CLASS_SPECS
            }
            records.append(
                GHSRecord(
                    source_file_id=source_file_id,
                    source_sheet=sheet.title,
                    source_row=row_index,
                    substance_name=values[positions["物質名称"]],
                    english_name="",
                    cas_text=values[positions["CAS RN"]],
                    cas_rns=tuple(extract_cas_rns(values[positions["CAS RN"]])),
                    ghs_result_id=values[positions["GHS分類結果_ID"]],
                    hazard_classes=hazard_classes,
                    model_label_url=clean_url(
                        values[positions["モデルラベル掲載ページURL"]]
                    ),
                    model_sds_url=clean_url(
                        values[positions["モデルＳＤＳ掲載ページURL"]]
                    ),
                )
            )

        connection.execute(
            """
            insert into raw_ghs_rows (
                source_file_id,
                source_sheet,
                row_number,
                row_kind,
                cells_json,
                parsed_ok
            ) values (?, ?, ?, ?, ?, ?)
            """,
            (
                source_file_id,
                sheet.title,
                row_index,
                row_kind,
                json.dumps(values, ensure_ascii=False),
                parsed_ok,
            ),
        )
        raw_row_count += 1

    return raw_row_count, records


def normalize_legal_records(
    connection: sqlite3.Connection,
    canonicalizer: Canonicalizer,
    records: list[LegalRecord],
) -> None:
    for record in records:
        substance_id = canonicalizer.get_or_create_substance(
            substance_name=record.substance_name,
            english_name=record.english_name,
            cas_rns=record.cas_rns,
        )
        cursor = connection.execute(
            """
            insert into legal_obligations (
                substance_id,
                source_file_id,
                source_sheet,
                source_row,
                substance_name,
                english_name,
                cas_text,
                section_title,
                section_number,
                list_index,
                label_threshold,
                sds_threshold,
                remarks,
                source_list_effective_date,
                raw_effective_date
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                substance_id,
                record.source_file_id,
                record.source_sheet,
                record.source_row,
                record.substance_name,
                record.english_name or None,
                record.cas_text,
                record.section_title,
                record.section_number or None,
                record.list_index or None,
                record.label_threshold,
                record.sds_threshold,
                record.remarks or None,
                record.source_list_effective_date,
                record.raw_effective_date or None,
            ),
        )
        legal_obligation_id = require_lastrowid(cursor)
        for cas_rn in record.cas_rns:
            connection.execute(
                """
                insert into legal_obligation_cas (
                    legal_obligation_id,
                    cas_rn,
                    cas_normalized
                ) values (?, ?, ?)
                """,
                (legal_obligation_id, cas_rn, normalize_cas(cas_rn)),
            )


def normalize_ghs_records(
    connection: sqlite3.Connection,
    canonicalizer: Canonicalizer,
    records: list[GHSRecord],
    hazard_class_ids: dict[str, int],
) -> None:
    for record in records:
        substance_id = canonicalizer.get_or_create_substance(
            substance_name=record.substance_name,
            english_name=record.english_name,
            cas_rns=record.cas_rns,
        )
        cursor = connection.execute(
            """
            insert into ghs_entries (
                substance_id,
                source_file_id,
                source_sheet,
                source_row,
                substance_name,
                english_name,
                cas_text,
                ghs_result_id,
                model_label_url,
                model_sds_url
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                substance_id,
                record.source_file_id,
                record.source_sheet,
                record.source_row,
                record.substance_name,
                record.english_name or None,
                record.cas_text,
                record.ghs_result_id,
                record.model_label_url,
                record.model_sds_url,
            ),
        )
        ghs_entry_id = require_lastrowid(cursor)

        for cas_rn in record.cas_rns:
            connection.execute(
                """
                insert into ghs_entry_cas (
                    ghs_entry_id,
                    cas_rn,
                    cas_normalized
                ) values (?, ?, ?)
                """,
                (ghs_entry_id, cas_rn, normalize_cas(cas_rn)),
            )

        for _, label in GHS_CLASS_SPECS:
            raw_result = record.hazard_classes[label]
            connection.execute(
                """
                insert into ghs_classifications (
                    ghs_entry_id,
                    hazard_class_id,
                    raw_result,
                    classification_state,
                    category_code,
                    is_assigned
                ) values (?, ?, ?, ?, ?, ?)
                """,
                (
                    ghs_entry_id,
                    hazard_class_ids[label],
                    raw_result,
                    classification_state(raw_result),
                    extract_category_codes(raw_result),
                    int(classification_state(raw_result) == "assigned"),
                ),
            )

        for pictogram_code in derive_pictograms(record.hazard_classes):
            connection.execute(
                """
                insert into ghs_pictograms (ghs_entry_id, pictogram_code)
                values (?, ?)
                """,
                (ghs_entry_id, pictogram_code),
            )


def load_reviewed_synonyms(canonicalizer: Canonicalizer, path: Path) -> None:
    if not path.exists():
        return
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            alias_raw = (row.get("alias_raw") or "").strip()
            if not alias_raw:
                continue
            substance_id = canonicalizer.resolve_substance_id(
                canonical_name=(row.get("canonical_name") or "").strip(),
                canonical_english_name=(
                    row.get("canonical_english_name") or ""
                ).strip(),
                canonical_cas=(row.get("canonical_cas") or "").strip(),
            )
            if substance_id is None:
                raise RuntimeError(
                    "could not resolve reviewed synonym row for "
                    f"{row.get('canonical_name') or row.get('canonical_english_name')}"
                )
            canonicalizer.add_alias(
                substance_id=substance_id,
                alias_raw=alias_raw,
                alias_type=(row.get("alias_type") or "reviewed_synonym").strip(),
                confidence=(row.get("confidence") or "high").strip(),
                exact_match_allowed=False,
            )


def seed_ghs_hazard_classes(connection: sqlite3.Connection) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for sort_order, (code, label_ja) in enumerate(GHS_CLASS_SPECS, start=1):
        cursor = connection.execute(
            """
            insert into ghs_hazard_classes (code, label_ja, sort_order)
            values (?, ?, ?)
            """,
            (code, label_ja, sort_order),
        )
        mapping[label_ja] = require_lastrowid(cursor)
    return mapping


def validate_snapshot(
    connection: sqlite3.Connection,
    *,
    source_snapshot_sha: str,
    expected_source_count: int,
    expected_ghs_classes_per_entry: int,
) -> None:
    source_count = scalar(connection, "select count(*) from source_files")
    raw_legal_row_total = scalar(connection, "select count(*) from raw_legal_rows")
    raw_legal_data_count = scalar(
        connection,
        "select count(*) from raw_legal_rows where row_kind = 'data' and parsed_ok = 1",
    )
    raw_ghs_row_total = scalar(connection, "select count(*) from raw_ghs_rows")
    raw_ghs_data_count = scalar(
        connection,
        "select count(*) from raw_ghs_rows where row_kind = 'data' and parsed_ok = 1",
    )
    legal_count = scalar(connection, "select count(*) from legal_obligations")
    legal_cas_count = scalar(connection, "select count(*) from legal_obligation_cas")
    ghs_count = scalar(connection, "select count(*) from ghs_entries")
    ghs_cas_count = scalar(connection, "select count(*) from ghs_entry_cas")
    hazard_class_count = scalar(connection, "select count(*) from ghs_hazard_classes")
    ghs_classification_count = scalar(
        connection, "select count(*) from ghs_classifications"
    )
    placeholder_url_count = scalar(
        connection,
        """
        select count(*)
        from ghs_entries
        where model_label_url = '-' or model_sds_url = '-'
        """,
    )

    if source_count != expected_source_count:
        raise RuntimeError(
            f"expected {expected_source_count} source files, found {source_count}"
        )
    if raw_legal_data_count != legal_count:
        raise RuntimeError(
            "raw legal data rows and normalized legal obligations diverged"
        )
    if raw_ghs_data_count != ghs_count:
        raise RuntimeError("raw GHS data rows and normalized GHS entries diverged")
    if ghs_classification_count != ghs_count * expected_ghs_classes_per_entry:
        raise RuntimeError(
            "GHS classification count does not match the hazard-class dimension size"
        )
    if placeholder_url_count != 0:
        raise RuntimeError("placeholder model label URLs were not normalized to NULL")
    if hazard_class_count != expected_ghs_classes_per_entry:
        raise RuntimeError(
            "GHS hazard-class dimension size does not match the expected schema"
        )

    baseline = SNAPSHOT_BASELINES.get(source_snapshot_sha)
    if baseline is None:
        return

    actual_counts = {
        "raw_legal_rows_total": raw_legal_row_total,
        "raw_legal_rows_data": raw_legal_data_count,
        "raw_ghs_rows_total": raw_ghs_row_total,
        "raw_ghs_rows_data": raw_ghs_data_count,
        "legal_obligations": legal_count,
        "legal_obligation_cas": legal_cas_count,
        "ghs_entries": ghs_count,
        "ghs_entry_cas": ghs_cas_count,
        "ghs_hazard_classes": hazard_class_count,
    }
    for key, expected_value in baseline.items():
        actual_value = actual_counts[key]
        if actual_value != expected_value:
            raise RuntimeError(
                f"snapshot baseline mismatch for {key}: expected {expected_value}, found {actual_value}"
            )


def write_build_meta(
    connection: sqlite3.Connection,
    *,
    manifest_sha: str,
    source_file_count: int,
    raw_legal_row_count: int,
    raw_ghs_row_count: int,
    fts_enabled: bool,
) -> None:
    connection.execute(
        """
        insert into build_meta (
            id,
            schema_version,
            built_at,
            source_manifest_sha256,
            source_file_count,
            raw_legal_row_count,
            raw_ghs_row_count,
            legal_obligation_count,
            ghs_entry_count,
            substance_count,
            fts_enabled
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            1,
            SCHEMA_VERSION,
            datetime.now(timezone.utc).isoformat(),
            manifest_sha,
            source_file_count,
            raw_legal_row_count,
            raw_ghs_row_count,
            scalar(connection, "select count(*) from legal_obligations"),
            scalar(connection, "select count(*) from ghs_entries"),
            scalar(connection, "select count(*) from substances"),
            int(fts_enabled),
        ),
    )


def compute_source_snapshot_sha(sources: list[dict[str, str]]) -> str:
    stable_sources = [
        {
            "filename": source["filename"],
            "key": source["key"],
            "sha256": source["sha256"],
            "url": source["url"],
        }
        for source in sorted(sources, key=lambda item: item["key"])
    ]
    payload = json.dumps(
        stable_sources,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def scalar(connection: sqlite3.Connection, query: str) -> int:
    return int(connection.execute(query).fetchone()[0])


def infer_sheet_effective_date(sheet_title: str) -> str | None:
    normalized = normalize_cell(sheet_title)
    match = re.search(r"R(\d+)\.(\d+)\.(\d+)", normalized)
    if match:
        reiwa_year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))
        return str(date(reiwa_year + 2018, month, day))
    match = re.search(r"R(\d+)\.(\d+)月", normalized)
    if match:
        reiwa_year = int(match.group(1))
        month = int(match.group(2))
        return str(date(reiwa_year + 2018, month, 1))
    return None


def detect_legal_header(values: list[str]) -> dict[str, int] | None:
    mapping = {}
    for index, value in enumerate(values):
        if "名称" in value and "英語名称" not in value:
            mapping["name"] = index
        if "英語名称" in value:
            mapping["english_name"] = index
        if "CAS RN" in value:
            mapping["cas"] = index
        if "ラベル表示に係る裾切値" in value:
            mapping["label_threshold"] = index
        if "SDS交付等に係る裾切値" in value or "ＳＤＳ交付等に係る裾切値" in value:
            mapping["sds_threshold"] = index
        if "番号" in value:
            if index == 0:
                mapping["list_index"] = index
            else:
                mapping["section_number"] = index
        if "備考" in value:
            mapping["remarks"] = index
        if "施行日" in value:
            mapping["effective_date"] = index
    required = {"name", "cas", "label_threshold", "sds_threshold"}
    if required.issubset(mapping):
        mapping.setdefault("list_index", 0)
        return mapping
    return None


def is_section_row(values: list[str]) -> bool:
    non_empty = [value for value in values if value]
    return (
        len(non_empty) == 1
        and "名称" not in non_empty[0]
        and "労働安全衛生" in non_empty[0]
    )


def is_legal_data_row(values: list[str], header_map: dict[str, int]) -> bool:
    name = values[header_map["name"]]
    cas = values[header_map["cas"]]
    label = values[header_map["label_threshold"]]
    sds = values[header_map["sds_threshold"]]
    return bool(name and cas and (label or sds))


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_url(value: str) -> str | None:
    cleaned = normalize_cell(value)
    if cleaned in {"", "-", "ー", "―"}:
        return None
    return cleaned


def first_non_empty(values: Iterable[str]) -> str:
    for value in values:
        if value:
            return value
    return ""


def _alias_confidence(value: str) -> str:
    normalized = normalize_synonym_text(value)
    if normalized.isascii() and 2 <= len(normalized) <= 6:
        return "high"
    return "medium"


def _should_keep_alias(alias_raw: str, alias_normalized: str) -> bool:
    if not alias_normalized:
        return False
    if len(alias_normalized) < 3 and not extract_cas_rns(alias_raw):
        return False
    return True


if __name__ == "__main__":
    main()
